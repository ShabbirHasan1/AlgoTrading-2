# ============================================================================
# Piotroski f score implementation (data scraped from yahoo finance)
# Author - Mayank Rasu

# Please report bugs/issues in the Q&A section
# =============================================================================


import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
tickers = ["ADANIPORTS.NS",	"ASIANPAINT.NS",	"AXISBANK.NS",	"BAJAJ-AUTO.NS",	"BAJFINANCE.NS",	"BAJAJFINSV.NS",	"BPCL.NS",	"BHARTIARTL.NS",	"INFRATEL.NS",	"BRITANNIA.NS",	"CIPLA.NS",	"COALINDIA.NS",	"DRREDDY.NS",	"EICHERMOT.NS",	"GAIL.NS",	"GRASIM.NS",	"HCLTECH.NS",	"HDFCBANK.NS",	"HEROMOTOCO.NS",	"HINDALCO.NS",	"HINDUNILVR.NS",	"HDFC.NS",	"ICICIBANK.NS",	"ITC.NS",	"IOC.NS",	"INDUSINDBK.NS",	"INFY.NS",	"JSWSTEEL.NS",	"KOTAKBANK.NS",	"LT.NS",	"M&M.NS",	"MARUTI.NS",	"NTPC.NS",	"NESTLEIND.NS",	"ONGC.NS",	"POWERGRID.NS",	"RELIANCE.NS",	"SHREECEM.NS",	"SBIN.NS",	"SUNPHARMA.NS",	"TCS.NS",	"TATAMOTORS.NS",	"TATASTEEL.NS",	"TECHM.NS",	"TITAN.NS",	"UPL.NS",	"ULTRACEMCO.NS",	"VEDL.NS",	"WIPRO.NS",	"ZEEL.NS",
           "ACC.NS",	"ABBOTINDIA.NS",	"ADANITRANS.NS",	"AMBUJACEM.NS",	"AUROPHARMA.NS",	"DMART.NS",	"BAJAJHLDNG.NS",	"BANDHANBNK.NS",	"BANKBARODA.NS",	"BERGEPAINT.NS",	"BIOCON.NS",	"BOSCHLTD.NS",	"CADILAHC.NS",	"COLPAL.NS",	"CONCOR.NS",	"DLF.NS",	"DABUR.NS",	"DIVISLAB.NS",	"GICRE.NS",	"GODREJCP.NS",	"HDFCAMC.NS",	"HDFCLIFE.NS",	"HAVELLS.NS",	"HINDPETRO.NS",	"HINDZINC.NS",	"ICICIGI.NS",	"ICICIPRULI.NS",	"IGL.NS",	"NAUKRI.NS",	"INDIGO.NS",	"LUPIN.NS",	"MARICO.NS",	"MOTHERSUMI.NS",	"MUTHOOTFIN.NS",	"NHPC.NS",	"NMDC.NS",	"OFSS.NS",	"PAGEIND.NS",	"PETRONET.NS",	"PIDILITIND.NS",	"PEL.NS",	"PFC.NS",	"PGHH.NS",	"PNB.NS",	"SBILIFE.NS",	"SRTRANSFIN.NS",	"SIEMENS.NS",	"TORNTPHARM.NS",	"UBL.NS",	"MCDOWELL-N.NS"]


#list of tickers whose financial data needs to be extracted
financial_dir_cy = {} #directory to store current year's information
financial_dir_py = {} #directory to store last year's information
financial_dir_py2 = {} #directory to store last to last year's information

for ticker in tickers:
    try:
        print("scraping financial statement data for ",ticker)
        temp_dir = {}
        temp_dir2 = {}
        temp_dir3 = {}
    #getting balance sheet data from yahoo finance for the given ticker
        url = 'https://in.finance.yahoo.com/quote/'+ticker+'/balance-sheet?p='+ticker
        page = requests.get(url)
        page_content = page.content
        soup = BeautifulSoup(page_content,'html.parser')
        tabl = soup.find_all("div", {"class" : "M(0) Whs(n) BdEnd Bdc($seperatorColor) D(itb)"})
        for t in tabl:
            rows = t.find_all("div", {"class" : "rw-expnded"})
            for row in rows:
                temp_dir[row.get_text(separator='|').split("|")[0]]=row.get_text(separator='|').split("|")[1]
                temp_dir2[row.get_text(separator='|').split("|")[0]]=row.get_text(separator='|').split("|")[2]
                temp_dir3[row.get_text(separator='|').split("|")[0]]=row.get_text(separator='|').split("|")[3]
        
        #getting income statement data from yahoo finance for the given ticker
        url = 'https://in.finance.yahoo.com/quote/'+ticker+'/financials?p='+ticker
        page = requests.get(url)
        page_content = page.content
        soup = BeautifulSoup(page_content,'html.parser')
        tabl = soup.find_all("div", {"class" : "M(0) Whs(n) BdEnd Bdc($seperatorColor) D(itb)"})
        for t in tabl:
            rows = t.find_all("div", {"class" : "rw-expnded"})
            for row in rows:
                temp_dir[row.get_text(separator='|').split("|")[0]]=row.get_text(separator='|').split("|")[2]
                temp_dir2[row.get_text(separator='|').split("|")[0]]=row.get_text(separator='|').split("|")[3]
                temp_dir3[row.get_text(separator='|').split("|")[0]]=row.get_text(separator='|').split("|")[4]
        
        #getting cashflow statement data from yahoo finance for the given ticker
        url = 'https://in.finance.yahoo.com/quote/'+ticker+'/cash-flow?p='+ticker
        page = requests.get(url)
        page_content = page.content
        soup = BeautifulSoup(page_content,'html.parser')
        tabl = soup.find_all("div", {"class" : "M(0) Whs(n) BdEnd Bdc($seperatorColor) D(itb)"})
        for t in tabl:
            rows = t.find_all("div", {"class" : "rw-expnded"})
            for row in rows:
                temp_dir[row.get_text(separator='|').split("|")[0]]=row.get_text(separator='|').split("|")[2]
                temp_dir2[row.get_text(separator='|').split("|")[0]]=row.get_text(separator='|').split("|")[3]
                temp_dir3[row.get_text(separator='|').split("|")[0]]=row.get_text(separator='|').split("|")[4] 
        
        #combining all extracted information with the corresponding ticker
        financial_dir_cy[ticker] = temp_dir
        financial_dir_py[ticker] = temp_dir2
        financial_dir_py2[ticker] = temp_dir3
    except:
        print("Problem scraping data for ",ticker)


#storing information in pandas dataframe
combined_financials_cy = pd.DataFrame(financial_dir_cy)
#combined_financials_cy.dropna(axis=1,inplace=True) #dropping columns with NaN values
combined_financials_py = pd.DataFrame(financial_dir_py)
#combined_financials_py.dropna(axis=1,inplace=True)
combined_financials_py2 = pd.DataFrame(financial_dir_py2)
#combined_financials_py2.dropna(axis=1,inplace=True)
tickers = combined_financials_cy.columns #updating the tickers list based on only those tickers whose values were successfully extracted

# selecting relevant financial information for each stock using fundamental data
stats = ["Net income available to common shareholders",
         "Total assets",
         "Net cash provided by operating activities",
         "Long-term debt",
         "Other long-term liabilities",
         "Total current assets",
         "Total current liabilities",
         "Common stock",
         "Total revenue",
         "Gross profit"] # change as required

indx = ["NetIncome","TotAssets","CashFlowOps","LTDebt","OtherLTDebt",
        "CurrAssets","CurrLiab","CommStock","TotRevenue","GrossProfit"]


def info_filter(df,stats,indx):
    """function to filter relevant financial information for each 
       stock and transforming string inputs to numeric"""
    tickers = df.columns
    all_stats = {}
    for ticker in tickers:
        try:
            temp = df[ticker]
            ticker_stats = []
            for stat in stats:
                ticker_stats.append(temp.loc[stat])
            all_stats['{}'.format(ticker)] = ticker_stats
        except:
            print("can't read data for ",ticker)
    
    all_stats_df = pd.DataFrame(all_stats,index=indx)
    
    # cleansing of fundamental data imported in dataframe
    all_stats_df[tickers] = all_stats_df[tickers].replace({',': ''}, regex=True)
    for ticker in all_stats_df.columns:
        all_stats_df[ticker] = pd.to_numeric(all_stats_df[ticker].values,errors='coerce')
    return all_stats_df

def piotroski_f(df_cy,df_py,df_py2):
    """function to calculate f score of each stock and output information as dataframe"""
    f_score = {}
    tickers = df_cy.columns
    for ticker in tickers:
        ROA_FS = int(df_cy.loc["NetIncome",ticker]/((df_cy.loc["TotAssets",ticker]+df_py.loc["TotAssets",ticker])/2) > 0)
        CFO_FS = int(df_cy.loc["CashFlowOps",ticker] > 0)
        ROA_D_FS = int(df_cy.loc["NetIncome",ticker]/(df_cy.loc["TotAssets",ticker]+df_py.loc["TotAssets",ticker])/2 > df_py.loc["NetIncome",ticker]/(df_py.loc["TotAssets",ticker]+df_py2.loc["TotAssets",ticker])/2)
        CFO_ROA_FS = int(df_cy.loc["CashFlowOps",ticker]/df_cy.loc["TotAssets",ticker] > df_cy.loc["NetIncome",ticker]/((df_cy.loc["TotAssets",ticker]+df_py.loc["TotAssets",ticker])/2))
        LTD_FS = int((df_cy.loc["LTDebt",ticker] + df_cy.loc["OtherLTDebt",ticker])<(df_py.loc["LTDebt",ticker] + df_py.loc["OtherLTDebt",ticker]))
        CR_FS = int((df_cy.loc["CurrAssets",ticker]/df_cy.loc["CurrLiab",ticker])>(df_py.loc["CurrAssets",ticker]/df_py.loc["CurrLiab",ticker]))
        DILUTION_FS = int(df_cy.loc["CommStock",ticker] <= df_py.loc["CommStock",ticker])
        GM_FS = int((df_cy.loc["GrossProfit",ticker]/df_cy.loc["TotRevenue",ticker])>(df_py.loc["GrossProfit",ticker]/df_py.loc["TotRevenue",ticker]))
        ATO_FS = int(df_cy.loc["TotRevenue",ticker]/((df_cy.loc["TotAssets",ticker]+df_py.loc["TotAssets",ticker])/2)>df_py.loc["TotRevenue",ticker]/((df_py.loc["TotAssets",ticker]+df_py2.loc["TotAssets",ticker])/2))
        f_score[ticker] = [ROA_FS,CFO_FS,ROA_D_FS,CFO_ROA_FS,LTD_FS,CR_FS,DILUTION_FS,GM_FS,ATO_FS]
    f_score_df = pd.DataFrame(f_score,index=["PosROA","PosCFO","ROAChange","Accruals","Leverage","Liquidity","Dilution","GM","ATO"])
    return f_score_df

# Selecting stocks with highest Piotroski f score
transformed_df_cy = info_filter(combined_financials_cy,stats,indx)
transformed_df_py = info_filter(combined_financials_py,stats,indx)
transformed_df_py2 = info_filter(combined_financials_py2,stats,indx)
final = pd.DataFrame()
f_score_df = piotroski_f(transformed_df_cy,transformed_df_py,transformed_df_py2)
final = f_score_df.sum().sort_values(ascending=False)
f = open("pitroski_f.txt", "w+")
f.write(str(final))
f.close()



wb = load_workbook("pitroski_f-Nifty100.xlsx")
writer = pd.ExcelWriter('pitroski_f-Nifty100.xlsx', engine='openpyxl') 
ws1 = wb.create_sheet("Pitroski_F_raw", 0) 
writer.book = wb
writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
f_score_df.to_excel(writer, "Pitroski_F_raw")
final.to_excel(writer,"Final_ranked")
writer.save()
#for r in dataframe_to_rows(f_score_df, index=True, header=True):
#   ws1.append(r)
#wb.save("pitroski_f-Nifty100.xlsx")



